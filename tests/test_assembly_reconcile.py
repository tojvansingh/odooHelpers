"""Offline reconciliation of the STOCK assembly against the sample plan.

The sample's forecast columns are stale/inconsistent (XLOOKUP cached against shifting
sheet state), so they are NOT reconciled here. The stock pull is clean: every Dish Towels
row in the plan's F-I columns must come straight from the All Stock 3.25 snapshot.
"""

from pathlib import Path

import openpyxl

SAMPLE = Path(__file__).resolve().parents[1] / "All Products 4.3.xlsx"


def _num(v) -> float:
    return v if isinstance(v, (int, float)) else 0


def test_stock_assembly_matches_plan():
    wb = openpyxl.load_workbook(SAMPLE, data_only=True)
    plan, stockws = wb["4.2 DT Plan"], wb["All Stock 3.25"]

    # All Stock 3.25 (Dish Towels) keyed by display name: (free, incoming, outgoing, on_hand)
    stock: dict[str, tuple] = {}
    for r in range(2, stockws.max_row + 1):
        if stockws.cell(r, 2).value != "Dish Towels":
            continue
        disp = stockws.cell(r, 5).value
        if disp:
            stock[disp] = (
                _num(stockws.cell(r, 6).value), _num(stockws.cell(r, 7).value),
                _num(stockws.cell(r, 8).value), _num(stockws.cell(r, 9).value),
            )

    checked = mismatches = 0
    for r in range(4, plan.max_row + 1):
        on_hand = plan.cell(r, 9).value
        if not isinstance(on_hand, (int, float)):
            continue
        checked += 1
        expected = (
            _num(plan.cell(r, 6).value), _num(plan.cell(r, 7).value),
            _num(plan.cell(r, 8).value), _num(on_hand),
        )
        if stock.get(plan.cell(r, 5).value) != expected:
            mismatches += 1

    assert checked == 334
    assert mismatches == 0
