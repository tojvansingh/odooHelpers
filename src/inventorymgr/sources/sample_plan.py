"""Read engine inputs from the user's sample Google-Sheets export ("4.2 DT Plan").

This source exists to reconcile the engine against the hand-built plan. In step 2
it is replaced by an Odoo source that assembles the same PlanInput objects live.
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

from ..model import PlanInput, Product

PLAN_TAB = "4.2 DT Plan"
DATA_START_ROW = 4
# AA..AK = projected demand months (Apr 2025 .. Feb 2026), 11 monthly columns.
FORECAST_COLS = range(27, 38)
# The sample assumes every existing incoming PO lands in one month ("Jul" = the
# 3rd projection step, index 2). The Odoo source will instead use real arrival dates.
INCOMING_ARRIVAL_INDEX = 2
# L1 in the sample = the Dish Towels MOQ / rounding increment.
MOQ_STEP = 50


def _num(v) -> float:
    return v if isinstance(v, (int, float)) else 0


def load_sample_plan(path: str | Path) -> list[tuple[int, PlanInput, float]]:
    """Return (sheet_row, PlanInput, sheet_order_qty) for each product row."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb[PLAN_TAB]
    rows: list[tuple[int, PlanInput, float]] = []
    for r in range(DATA_START_ROW, ws.max_row + 1):
        on_hand = ws.cell(r, 9).value
        if not isinstance(on_hand, (int, float)):
            continue
        product = Product(
            name=ws.cell(r, 1).value,
            display_name=ws.cell(r, 5).value,
            class_name=ws.cell(r, 2).value,
            collection=ws.cell(r, 3).value,
            on_hand=_num(on_hand),
            incoming=_num(ws.cell(r, 7).value),
            outgoing=_num(ws.cell(r, 8).value),
            average_cost=_num(ws.cell(r, 4).value),
        )
        forecasts = [_num(ws.cell(r, c).value) for c in FORECAST_COLS]
        pi = PlanInput(product, forecasts, {INCOMING_ARRIVAL_INDEX: product.incoming})
        sheet_order = _num(ws.cell(r, 12).value)
        rows.append((r, pi, sheet_order))
    return rows
